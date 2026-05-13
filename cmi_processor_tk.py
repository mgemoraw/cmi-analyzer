"""
Excel Batch Processor GUI
=========================

FEATURES
--------
✔ Modern Tkinter GUI
✔ Select Input Folder
✔ Select Output Folder
✔ Select Config File
✔ Progress Bar
✔ Real-time Logs
✔ Start Button
✔ Thread-safe Processing
✔ Error Handling
✔ Clean OOP Architecture

------------------------------------------------------------
INSTALLATION
------------------------------------------------------------

pip install pandas openpyxl tqdm pyyaml

------------------------------------------------------------
RUN
------------------------------------------------------------

python excel_gui.py
"""

import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from threading import Thread
from pathlib import Path
import queue
import yaml
import logging
import pandas as pd
from concurrent.futures import ThreadPoolExecutor
from math import ceil

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill


# ============================================================
# CONFIG MANAGER
# ============================================================

class ConfigManager:

    def __init__(self, config_path):
        self.config_path = Path(config_path)

    def load(self):

        if not self.config_path.exists():
            raise FileNotFoundError(
                f"Config file not found: {self.config_path}"
            )

        with open(self.config_path, "r") as f:
            return yaml.safe_load(f)


# ============================================================
# LOGGER
# ============================================================

class GuiLogger:

    def __init__(self, log_queue):
        self.log_queue = log_queue

    def info(self, message):
        self.log_queue.put(f"[INFO] {message}")

    def warning(self, message):
        self.log_queue.put(f"[WARNING] {message}")

    def error(self, message):
        self.log_queue.put(f"[ERROR] {message}")


# ============================================================
# VALIDATOR
# ============================================================

class HeaderValidator:

    def __init__(self, required_headers):
        self.required_headers = required_headers

    def validate(self, df, sheet_name):

        required = self.required_headers.get(sheet_name, [])

        missing = []

        for col in required:
            if col not in df.columns:
                missing.append(col)

        return missing


# ============================================================
# EXCEL READER
# ============================================================

class ExcelReader:

    REQUIRED_SHEETS = [
        "equipment",
        "mpdm",
        "dailyvariables"
    ]

    def __init__(self, file_path, validator, logger):

        self.file_path = Path(file_path)
        self.validator = validator
        self.logger = logger

    def process(self):

        result = {}

        try:

            excel = pd.ExcelFile(self.file_path)

            sheet_map = {
                s.lower(): s
                for s in excel.sheet_names
            }

            for sheet in self.REQUIRED_SHEETS:

                if sheet not in sheet_map:

                    self.logger.warning(
                        f"{self.file_path.name} missing '{sheet}' sheet"
                    )

                    result[sheet] = pd.DataFrame()
                    continue

                actual_name = sheet_map[sheet]

                df = pd.read_excel(
                    self.file_path,
                    sheet_name=actual_name,
                    engine="openpyxl"
                )

                df = df.dropna(how="all")

                missing_headers = self.validator.validate(df, sheet)

                if missing_headers:
                    self.logger.warning(
                        f"{self.file_path.name} -> "
                        f"{sheet} missing headers: {missing_headers}"
                    )

                result[sheet] = df

        except Exception as e:

            self.logger.error(
                f"Error processing {self.file_path.name}: {e}"
            )

        return result


# ============================================================
# COLLECTOR
# ============================================================

class DataCollector:

    def __init__(self):

        self.equipment = []
        self.mpdm = []
        self.dailyvariables = []

    def add(self, data):

        if not data["equipment"].empty:
            self.equipment.append(data["equipment"])

        if not data["mpdm"].empty:
            self.mpdm.append(data["mpdm"])

        if not data["dailyvariables"].empty:
            self.dailyvariables.append(data["dailyvariables"])

    def combine(self):

        equipment_df = pd.concat(
            self.equipment,
            ignore_index=True
        ) if self.equipment else pd.DataFrame()

        mpdm_df = pd.concat(
            self.mpdm,
            ignore_index=True
        ) if self.mpdm else pd.DataFrame()

        daily_df = pd.concat(
            self.dailyvariables,
            ignore_index=True
        ) if self.dailyvariables else pd.DataFrame()

        return equipment_df, mpdm_df, daily_df


# ============================================================
# FORMATTER
# ============================================================

class ExcelFormatter:

    HEADER_FILL = PatternFill(
        start_color="1F4E78",
        end_color="1F4E78",
        fill_type="solid"
    )

    HEADER_FONT = Font(
        bold=True,
        color="FFFFFF"
    )

    @staticmethod
    def format_workbook(file_path):

        wb = load_workbook(file_path)

        for ws in wb.worksheets:

            for cell in ws[1]:
                cell.font = ExcelFormatter.HEADER_FONT
                cell.fill = ExcelFormatter.HEADER_FILL

            for column_cells in ws.columns:

                length = max(
                    len(str(cell.value))
                    if cell.value else 0
                    for cell in column_cells
                )

                ws.column_dimensions[
                    column_cells[0].column_letter
                ].width = min(length + 5, 50)

        wb.save(file_path)


# ============================================================
# WRITER
# ============================================================

class ExcelWriter:

    def __init__(self, output_dir, logger):

        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(exist_ok=True)

        self.logger = logger

    def save_dailyvariables(self, df):

        file_path = self.output_dir / "all_dailyvariables.xlsx"

        with pd.ExcelWriter(
                file_path,
                engine="openpyxl"
        ) as writer:

            df.to_excel(
                writer,
                sheet_name="dailyvariables",
                index=False
            )

        ExcelFormatter.format_workbook(file_path)

        self.logger.info("Saved all_dailyvariables.xlsx")

    def save_chunks(
            self,
            equipment_df,
            mpdm_df,
            chunk_size
    ):

        max_rows = max(
            len(equipment_df),
            len(mpdm_df)
        )

        total_files = ceil(max_rows / chunk_size)

        for i in range(total_files):

            start = i * chunk_size
            end = start + chunk_size

            eq_chunk = equipment_df.iloc[start:end]
            mpdm_chunk = mpdm_df.iloc[start:end]

            file_name = f"dataset_chunk_{i + 1}.xlsx"

            file_path = self.output_dir / file_name

            with pd.ExcelWriter(
                    file_path,
                    engine="openpyxl"
            ) as writer:

                eq_chunk.to_excel(
                    writer,
                    sheet_name="equipment",
                    index=False
                )

                mpdm_chunk.to_excel(
                    writer,
                    sheet_name="mpdm",
                    index=False
                )

            ExcelFormatter.format_workbook(file_path)

            self.logger.info(f"Saved {file_name}")


# ============================================================
# PROCESSOR
# ============================================================

class ExcelBatchProcessor:

    def __init__(
            self,
            input_dir,
            output_dir,
            config_path,
            logger,
            progress_callback
    ):

        self.input_dir = Path(input_dir)

        config = ConfigManager(config_path).load()

        self.chunk_size = config["chunk_size"]
        self.max_workers = config["max_workers"]

        self.logger = logger

        self.validator = HeaderValidator(
            config["required_headers"]
        )

        self.collector = DataCollector()

        self.writer = ExcelWriter(
            output_dir,
            logger
        )

        self.progress_callback = progress_callback

    def process_single_file(self, file_path):

        reader = ExcelReader(
            file_path,
            self.validator,
            self.logger
        )

        return reader.process()

    def run(self):

        files = list(
            self.input_dir.glob("*.xlsx")
        )

        if not files:
            raise Exception("No Excel files found")

        total = len(files)

        self.logger.info(f"Found {total} Excel files")

        with ThreadPoolExecutor(
                max_workers=self.max_workers
        ) as executor:

            for index, result in enumerate(
                    executor.map(
                        self.process_single_file,
                        files
                    )
            ):

                self.collector.add(result)

                progress = int(
                    ((index + 1) / total) * 100
                )

                self.progress_callback(progress)

        equipment_df, mpdm_df, daily_df = (
            self.collector.combine()
        )

        self.writer.save_dailyvariables(daily_df)

        self.writer.save_chunks(
            equipment_df,
            mpdm_df,
            self.chunk_size
        )

        self.logger.info("Processing completed")


# ============================================================
# GUI
# ============================================================

class ExcelProcessorGUI:

    def __init__(self, root):

        self.root = root

        self.root.title(
            "Excel Batch Processor"
        )

        self.root.geometry("850x600")

        self.log_queue = queue.Queue()

        self.create_widgets()

        self.update_logs()

    def create_widgets(self):

        title = tk.Label(
            self.root,
            text="Excel Batch Processor",
            font=("Arial", 18, "bold")
        )

        title.pack(pady=10)

        # INPUT FOLDER
        self.input_var = tk.StringVar()

        self.create_path_selector(
            "Input Folder",
            self.input_var,
            self.select_input_folder
        )

        # OUTPUT FOLDER
        self.output_var = tk.StringVar()

        self.create_path_selector(
            "Output Folder",
            self.output_var,
            self.select_output_folder
        )

        # CONFIG FILE
        self.config_var = tk.StringVar()

        self.create_path_selector(
            "Config File",
            self.config_var,
            self.select_config_file
        )

        # START BUTTON
        self.start_btn = tk.Button(
            self.root,
            text="Start Processing",
            font=("Arial", 12, "bold"),
            bg="#1F4E78",
            fg="white",
            command=self.start_processing
        )

        self.start_btn.pack(pady=10)

        # PROGRESS BAR
        self.progress = ttk.Progressbar(
            self.root,
            orient="horizontal",
            length=700,
            mode="determinate"
        )

        self.progress.pack(pady=10)

        # LOG AREA
        self.log_text = tk.Text(
            self.root,
            height=20
        )

        self.log_text.pack(
            padx=10,
            pady=10,
            fill="both",
            expand=True
        )

    def create_path_selector(
            self,
            label_text,
            variable,
            button_command
    ):

        frame = tk.Frame(self.root)
        frame.pack(fill="x", padx=10, pady=5)

        label = tk.Label(
            frame,
            text=label_text,
            width=15,
            anchor="w"
        )

        label.pack(side="left")

        entry = tk.Entry(
            frame,
            textvariable=variable
        )

        entry.pack(
            side="left",
            fill="x",
            expand=True,
            padx=5
        )

        button = tk.Button(
            frame,
            text="Browse",
            command=button_command
        )

        button.pack(side="right")

    def select_input_folder(self):

        folder = filedialog.askdirectory()

        if folder:
            self.input_var.set(folder)

    def select_output_folder(self):

        folder = filedialog.askdirectory()

        if folder:
            self.output_var.set(folder)

    def select_config_file(self):

        file = filedialog.askopenfilename(
            filetypes=[("YAML Files", "*.yaml")]
        )

        if file:
            self.config_var.set(file)

    def log(self, message):

        self.log_queue.put(message)

    def update_logs(self):

        while not self.log_queue.empty():

            msg = self.log_queue.get()

            self.log_text.insert(
                tk.END,
                msg + "\n"
            )

            self.log_text.see(tk.END)

        self.root.after(
            100,
            self.update_logs
        )

    def update_progress(self, value):

        self.progress["value"] = value

        self.root.update_idletasks()

    def start_processing(self):

        if not self.input_var.get():

            messagebox.showerror(
                "Error",
                "Select input folder"
            )

            return

        if not self.output_var.get():

            messagebox.showerror(
                "Error",
                "Select output folder"
            )

            return

        if not self.config_var.get():

            messagebox.showerror(
                "Error",
                "Select config file"
            )

            return

        self.start_btn.config(state="disabled")

        thread = Thread(
            target=self.process_files
        )

        thread.start()

    def process_files(self):

        try:

            logger = GuiLogger(
                self.log_queue
            )

            processor = ExcelBatchProcessor(
                input_dir=self.input_var.get(),
                output_dir=self.output_var.get(),
                config_path=self.config_var.get(),
                logger=logger,
                progress_callback=self.update_progress
            )

            processor.run()

            messagebox.showinfo(
                "Success",
                "Processing completed!"
            )

        except Exception as e:

            messagebox.showerror(
                "Error",
                str(e)
            )

        finally:

            self.start_btn.config(state="normal")


# ============================================================
# MAIN
# ============================================================

if __name__ == "__main__":

    root = tk.Tk()

    app = ExcelProcessorGUI(root)

    root.mainloop()