"""
Advanced Excel Batch Processor
==============================

FEATURES
--------
✔ Object-Oriented Design
✔ Reads multiple Excel files
✔ Case-insensitive sheet detection
✔ Validation of required headers
✔ Logging system
✔ Progress bar
✔ Configuration file support
✔ Automatic formatting
✔ Multi-threaded processing
✔ Removes empty rows
✔ Combines ALL dailyvariables into ONE file
✔ Combines equipment + mpdm into chunked Excel files
✔ equipment and mpdm saved in SAME workbook
✔ Each workbook contains:
      - equipment sheet
      - mpdm sheet
✔ Splits records into configurable chunk size
✔ Easy to maintain and extend

------------------------------------------------------------
INSTALLATION
------------------------------------------------------------

pip install pandas openpyxl tqdm pyyaml

------------------------------------------------------------
PROJECT STRUCTURE
------------------------------------------------------------

project/
│
├── input_files/
│      file1.xlsx
│      file2.xlsx
│
├── output/
│
├── config.yaml
│
└── advanced_excel_processor.py

------------------------------------------------------------
CONFIGURATION (config.yaml)
------------------------------------------------------------

chunk_size: 100
max_workers: 4

required_headers:
  equipment:
    - Date
    - Project Code

  mpdm:
    - Date
    - Project Code

  dailyvariables:
    - Date
    - Project Code

------------------------------------------------------------
RUN
------------------------------------------------------------

python advanced_excel_processor.py
"""

from pathlib import Path
from concurrent.futures import ThreadPoolExecutor
from math import ceil
import logging
import yaml
import pandas as pd

from tqdm import tqdm
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill


# ============================================================
# CONFIG MANAGER
# ============================================================

class ConfigManager:

    def __init__(self, config_path="config.yaml"):

        self.default_config = {
            "chunk_size": 100,
            "max_workers": 4,
            "required_headers": {
                "equipment": ["Date", "Project Code"],
                "mpdm": ["Date", "Project Code"],
                "dailyvariables": ["Date", "Project Code"]
            }
        }

        self.config_path = Path(config_path)

    def load(self):

        if not self.config_path.exists():
            self.create_default_config()

        with open(self.config_path, "r") as f:
            return yaml.safe_load(f)

    def create_default_config(self):

        with open(self.config_path, "w") as f:
            yaml.dump(self.default_config, f)

        print(f"Default config created: {self.config_path}")


# ============================================================
# LOGGER
# ============================================================

class LoggerManager:

    @staticmethod
    def setup_logger():

        Path("logs").mkdir(exist_ok=True)

        logging.basicConfig(
            filename="logs/process.log",
            level=logging.INFO,
            format="%(asctime)s - %(levelname)s - %(message)s"
        )

        return logging.getLogger("ExcelProcessor")


# ============================================================
# VALIDATOR
# ============================================================

class HeaderValidator:

    def __init__(self, required_headers):
        self.required_headers = required_headers

    def validate(self, df, sheet_name):

        required = self.required_headers.get(sheet_name, [])

        missing = []

        for col in required:
            if col not in df.columns:
                missing.append(col)

        return missing


# ============================================================
# EXCEL READER
# ============================================================

class ExcelReader:

    REQUIRED_SHEETS = [
        "equipment",
        "mpdm",
        "dailyvariables"
    ]

    def __init__(self, file_path, validator, logger):

        self.file_path = Path(file_path)
        self.validator = validator
        self.logger = logger

    def read_sheet(self, actual_sheet_name):

        try:

            df = pd.read_excel(
                self.file_path,
                sheet_name=actual_sheet_name,
                engine="openpyxl"
            )

            df = df.dropna(how="all")

            return df

        except Exception as e:

            self.logger.error(
                f"Error reading {actual_sheet_name} "
                f"in {self.file_path.name}: {e}"
            )

            return pd.DataFrame()

    def process(self):

        result = {}

        try:

            excel = pd.ExcelFile(self.file_path)

            sheet_map = {
                s.lower(): s
                for s in excel.sheet_names
            }

            for sheet in self.REQUIRED_SHEETS:

                if sheet.lower() not in sheet_map:

                    self.logger.warning(
                        f"Missing sheet '{sheet}' "
                        f"in {self.file_path.name}"
                    )

                    result[sheet] = pd.DataFrame()
                    continue

                actual_name = sheet_map[sheet.lower()]

                df = self.read_sheet(actual_name)

                missing_headers = self.validator.validate(df, sheet)

                if missing_headers:

                    self.logger.warning(
                        f"{self.file_path.name} -> "
                        f"{sheet} missing headers: {missing_headers}"
                    )

                result[sheet] = df

        except Exception as e:

            self.logger.error(
                f"Error processing {self.file_path.name}: {e}"
            )

        return result


# ============================================================
# DATA COLLECTOR
# ============================================================

class DataCollector:

    def __init__(self):

        self.equipment = []
        self.mpdm = []
        self.dailyvariables = []

    def add(self, data):

        if not data["equipment"].empty:
            self.equipment.append(data["equipment"])

        if not data["mpdm"].empty:
            self.mpdm.append(data["mpdm"])

        if not data["dailyvariables"].empty:
            self.dailyvariables.append(data["dailyvariables"])

    def combine(self):

        equipment_df = pd.concat(
            self.equipment,
            ignore_index=True
        ) if self.equipment else pd.DataFrame()

        mpdm_df = pd.concat(
            self.mpdm,
            ignore_index=True
        ) if self.mpdm else pd.DataFrame()

        daily_df = pd.concat(
            self.dailyvariables,
            ignore_index=True
        ) if self.dailyvariables else pd.DataFrame()

        return equipment_df, mpdm_df, daily_df


# ============================================================
# FORMATTER
# ============================================================

class ExcelFormatter:

    HEADER_FILL = PatternFill(
        start_color="1F4E78",
        end_color="1F4E78",
        fill_type="solid"
    )

    HEADER_FONT = Font(
        bold=True,
        color="FFFFFF"
    )

    @staticmethod
    def format_workbook(file_path):

        wb = load_workbook(file_path)

        for ws in wb.worksheets:

            for cell in ws[1]:

                cell.font = ExcelFormatter.HEADER_FONT
                cell.fill = ExcelFormatter.HEADER_FILL

            # Auto width
            for column_cells in ws.columns:

                length = max(
                    len(str(cell.value))
                    if cell.value is not None else 0
                    for cell in column_cells
                )

                ws.column_dimensions[
                    column_cells[0].column_letter
                ].width = min(length + 5, 50)

        wb.save(file_path)


# ============================================================
# EXCEL WRITER
# ============================================================

class ExcelWriter:

    def __init__(self, output_dir, logger):

        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(exist_ok=True)

        self.logger = logger

    def save_dailyvariables(self, df):

        file_path = self.output_dir / "all_dailyvariables.xlsx"

        with pd.ExcelWriter(
            file_path,
            engine="openpyxl"
        ) as writer:

            df.to_excel(
                writer,
                sheet_name="dailyvariables",
                index=False
            )

        ExcelFormatter.format_workbook(file_path)

        self.logger.info("Saved dailyvariables")

    def save_combined_chunks(
            self,
            equipment_df,
            mpdm_df,
            chunk_size
    ):

        max_rows = max(
            len(equipment_df),
            len(mpdm_df)
        )

        total_files = ceil(max_rows / chunk_size)

        for i in range(total_files):

            start = i * chunk_size
            end = start + chunk_size

            eq_chunk = equipment_df.iloc[start:end]
            mpdm_chunk = mpdm_df.iloc[start:end]

            file_name = f"dataset_chunk_{i + 1}.xlsx"

            file_path = self.output_dir / file_name

            with pd.ExcelWriter(
                    file_path,
                    engine="openpyxl"
            ) as writer:

                eq_chunk.to_excel(
                    writer,
                    sheet_name="equipment",
                    index=False
                )

                mpdm_chunk.to_excel(
                    writer,
                    sheet_name="mpdm",
                    index=False
                )

            ExcelFormatter.format_workbook(file_path)

            self.logger.info(f"Saved {file_name}")


# ============================================================
# MAIN PROCESSOR
# ============================================================

class ExcelBatchProcessor:

    def __init__(
            self,
            input_dir="input_files",
            output_dir="output",
            config_path="config.yaml"
    ):

        self.config = ConfigManager(config_path).load()

        self.logger = LoggerManager.setup_logger()

        self.collector = DataCollector()

        self.validator = HeaderValidator(
            self.config["required_headers"]
        )

        self.writer = ExcelWriter(
            output_dir,
            self.logger
        )

        self.input_dir = Path(input_dir)

    def process_single_file(self, file_path):

        reader = ExcelReader(
            file_path,
            self.validator,
            self.logger
        )

        return reader.process()

    def process(self):

        files = list(
            self.input_dir.glob("*.xlsx")
        )

        if not files:

            print("No Excel files found.")
            return

        self.logger.info(
            f"Found {len(files)} Excel files"
        )

        with ThreadPoolExecutor(
                max_workers=self.config["max_workers"]
        ) as executor:

            results = list(
                tqdm(
                    executor.map(
                        self.process_single_file,
                        files
                    ),
                    total=len(files),
                    desc="Processing Files"
                )
            )

        for result in results:
            self.collector.add(result)

        equipment_df, mpdm_df, daily_df = (
            self.collector.combine()
        )

        self.writer.save_dailyvariables(daily_df)

        self.writer.save_combined_chunks(
            equipment_df,
            mpdm_df,
            self.config["chunk_size"]
        )

        self.logger.info("Processing completed")

        print("\nProcessing completed successfully!")


# ============================================================
# MAIN
# ============================================================

if __name__ == "__main__":

    processor = ExcelBatchProcessor(
        input_dir="input_files",
        output_dir="output",
        config_path="config.yaml"
    )

    processor.process()