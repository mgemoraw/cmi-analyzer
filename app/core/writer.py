# app/core/excel_writer.py

from pathlib import Path
from math import ceil

import pandas as pd

from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from openpyxl.styles import Alignment


class ExcelWriter:
    """
    Excel Writer Class

    Responsibilities:
    -----------------
    1. Save dailyvariables
    2. Save chunked workbooks
    3. Format Excel files
    4. Auto-adjust columns
    """

    # =====================================================
    # INIT
    # =====================================================

    def __init__(
        self,
        output_dir,
        logger_callback=None
    ):

        self.output_dir = Path(output_dir)

        self.output_dir.mkdir(
            parents=True,
            exist_ok=True
        )

        self.logger_callback = logger_callback

    # =====================================================
    # LOGGER
    # =====================================================

    def log(self, message):

        print(message)

        if self.logger_callback:
            self.logger_callback(message)

    # =====================================================
    # FORMAT WORKBOOK
    # =====================================================

    def format_workbook(self, file_path):

        try:

            wb = load_workbook(file_path)

            for ws in wb.worksheets:

                # ==================================
                # HEADER STYLE
                # ==================================

                for cell in ws[1]:

                    cell.font = Font(
                        bold=True,
                        color="FFFFFF"
                    )

                    cell.fill = PatternFill(
                        start_color="1F4E78",
                        end_color="1F4E78",
                        fill_type="solid"
                    )

                    cell.alignment = Alignment(
                        horizontal="center",
                        vertical="center"
                    )

                # ==================================
                # AUTO WIDTH
                # ==================================

                for column in ws.columns:

                    max_length = 0

                    column_letter = (
                        column[0].column_letter
                    )

                    for cell in column:

                        try:

                            if cell.value:

                                max_length = max(
                                    max_length,
                                    len(
                                        str(cell.value)
                                    )
                                )

                        except:
                            pass

                    adjusted_width = min(
                        max_length + 5,
                        60
                    )

                    ws.column_dimensions[
                        column_letter
                    ].width = adjusted_width

                # ==================================
                # FREEZE HEADER
                # ==================================

                ws.freeze_panes = "A2"

            wb.save(file_path)

        except Exception as e:

            self.log(
                f"[ERROR] Formatting failed: "
                f"{str(e)}"
            )

    # =====================================================
    # SAVE DAILYVARIABLES
    # =====================================================

    def save_dailyvariables(
        self,
        dailyvariables_df
    ):

        if dailyvariables_df.empty:

            self.log(
                "[WARNING] No dailyvariables "
                "data to save"
            )

            return

        output_file = (
            self.output_dir
            / "all_dailyvariables.xlsx"
        )

        self.log(
            "[INFO] Saving dailyvariables..."
        )

        with pd.ExcelWriter(
            output_file,
            engine="openpyxl"
        ) as writer:

            dailyvariables_df.to_excel(
                writer,
                sheet_name="dailyvariables",
                index=False
            )

        self.format_workbook(output_file)

        self.log(
            "[SUCCESS] Saved "
            "all_dailyvariables.xlsx"
        )

    # =====================================================
    # SAVE CHUNK FILES
    # =====================================================

    def save_chunk_files(
        self,
        equipment_df,
        mpdm_df,
        chunk_size=100
    ):

        max_rows = max(
            len(equipment_df),
            len(mpdm_df)
        )

        if max_rows == 0:

            self.log(
                "[WARNING] No equipment/mpdm "
                "data to save"
            )

            return

        total_chunks = ceil(
            max_rows / chunk_size
        )

        self.log(
            f"[INFO] Creating "
            f"{total_chunks} chunk files..."
        )

        for i in range(total_chunks):

            start = i * chunk_size

            end = start + chunk_size

            equipment_chunk = (
                equipment_df.iloc[start:end]
            )

            mpdm_chunk = (
                mpdm_df.iloc[start:end]
            )

            output_file = (
                self.output_dir
                / f"dataset_chunk_{i+1}.xlsx"
            )

            with pd.ExcelWriter(
                output_file,
                engine="openpyxl"
            ) as writer:

                # ==========================
                # EQUIPMENT SHEET
                # ==========================

                equipment_chunk.to_excel(
                    writer,
                    sheet_name="equipment",
                    index=False
                )

                # ==========================
                # MPDM SHEET
                # ==========================

                mpdm_chunk.to_excel(
                    writer,
                    sheet_name="mpdm",
                    index=False
                )

            self.format_workbook(output_file)

            self.log(
                f"[SUCCESS] Saved "
                f"{output_file.name}"
            )

    # =====================================================
    # SAVE SINGLE DATAFRAME
    # =====================================================

    def save_dataframe(
        self,
        dataframe,
        file_name,
        sheet_name="Sheet1"
    ):

        output_file = (
            self.output_dir / file_name
        )

        with pd.ExcelWriter(
            output_file,
            engine="openpyxl"
        ) as writer:

            dataframe.to_excel(
                writer,
                sheet_name=sheet_name,
                index=False
            )

        self.format_workbook(output_file)

        self.log(
            f"[SUCCESS] Saved {file_name}"
        )