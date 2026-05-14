# app/core/excel_processor.py

from pathlib import Path
from math import ceil
from concurrent.futures import ThreadPoolExecutor

import pandas as pd

from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.styles import PatternFill


class ExcelProcessor:
    """
    Main Excel Processing Engine

    Responsibilities:
    -----------------
    1. Read all Excel files
    2. Extract:
        - equipment
        - mpdm
        - dailyvariables
    3. Validate sheets
    4. Combine dailyvariables
    5. Split equipment/mpdm into chunks
    6. Save output files
    7. Report progress/logs/statistics
    """

    REQUIRED_SHEETS = [
        "equipment",
        "mpdm",
        "dailyvariables"
    ]

    SHEET_ALIASES = {
        "equipment": [
            "equipment",
            "excavator",
            "dozer",
            "truck",
            "roller",
            "grader",
            "labor",
            "other",
        ],
        "mpdm": ["mpdm"],
        "worksampling": [
            "worksampling",],
        "dailyvariables": [
            "daily_variables",
            "daily_variables"
        ]
    }

    # =====================================================
    # INIT
    # =====================================================

    def __init__(
        self,
        input_dir,
        output_dir,
        chunk_size=100,
        max_workers=4,
        logger_callback=None,
        progress_callback=None,
        statistics_callback=None,
        current_file_callback=None,
    ):

        self.input_dir = Path(input_dir)

        self.output_dir = Path(output_dir)

        self.output_dir.mkdir(
            parents=True,
            exist_ok=True
        )

        self.chunk_size = chunk_size

        self.max_workers = max_workers

        # CALLBACKS
        self.logger_callback = logger_callback

        self.progress_callback = progress_callback

        self.statistics_callback = (
            statistics_callback
        )

        self.current_file_callback = (
            current_file_callback
        )

        # DATA STORAGE
        self.equipment_data = []

        self.mpdm_data = []

        self.dailyvariables_data = []

        # STATISTICS
        self.total_files = 0

        self.processed_files = 0

        self.total_equipment_rows = 0

        self.total_mpdm_rows = 0

        self.total_daily_rows = 0

        self.total_errors = 0

    # =====================================================
    # LOGGER
    # =====================================================

    def log(self, message):

        print(message)

        if self.logger_callback:
            self.logger_callback(message)

    # =====================================================
    # UPDATE PROGRESS
    # =====================================================

    def update_progress(self):

        if self.total_files == 0:
            return

        progress = int(
            (self.processed_files / self.total_files)
            * 100
        )

        if self.progress_callback:
            self.progress_callback(progress)

    # =====================================================
    # UPDATE STATISTICS
    # =====================================================

    def update_statistics(self):

        if self.statistics_callback:

            self.statistics_callback(
                files=self.processed_files,
                equipment=self.total_equipment_rows,
                mpdm=self.total_mpdm_rows,
                daily=self.total_daily_rows,
                errors=self.total_errors
            )

    # =====================================================
    # FIND SHEET
    # =====================================================

    def get_sheet_name(
        self,
        available_sheets,
        target_sheet_names
    ):

        if isinstance(target_sheet_names, str):
            target_sheet_names = [target_sheet_names]

        sheet_map = {
            s.lower(): s
            for s in available_sheets
        }

        for target in target_sheet_names:
            actual = sheet_map.get(target.lower())
            if actual:
                return actual

        return None

    # =====================================================
    # READ EXCEL FILE
    # =====================================================

    def process_single_file(self, file_path):

        try:

            if self.current_file_callback:

                self.current_file_callback(
                    file_path.name
                )

            self.log(
                f"[INFO] Processing: {file_path.name}"
            )

            excel = pd.ExcelFile(file_path)

            # ============================
            # EQUIPMENT
            # ============================

            equipment_sheet = self.get_sheet_name(
                excel.sheet_names,
                self.SHEET_ALIASES["equipment"]
            )

            if equipment_sheet:

                equipment_df = pd.read_excel(
                    file_path,
                    sheet_name=equipment_sheet,
                    engine="openpyxl"
                )

                equipment_df = (
                    equipment_df.dropna(
                        how="all"
                    )
                )

                if not equipment_df.empty:

                    self.equipment_data.append(
                        equipment_df
                    )

                    self.total_equipment_rows += (
                        len(equipment_df)
                    )

            else:

                self.log(
                    f"[WARNING] equipment sheet missing "
                    f"in {file_path.name}"
                )

            # ============================
            # MPDM
            # ============================

            mpdm_sheet = self.get_sheet_name(
                excel.sheet_names,
                self.SHEET_ALIASES["mpdm"]
            )

            if mpdm_sheet:

                mpdm_df = pd.read_excel(
                    file_path,
                    sheet_name=mpdm_sheet,
                    engine="openpyxl"
                )

                mpdm_df = mpdm_df.dropna(
                    how="all"
                )

                if not mpdm_df.empty:

                    self.mpdm_data.append(
                        mpdm_df
                    )

                    self.total_mpdm_rows += (
                        len(mpdm_df)
                    )

            else:

                self.log(
                    f"[WARNING] mpdm sheet missing "
                    f"in {file_path.name}"
                )

            # ============================
            # DAILYVARIABLES
            # ============================

            daily_sheet = self.get_sheet_name(
                excel.sheet_names,
                self.SHEET_ALIASES["daily_variables"]
            )

            if daily_sheet:

                daily_df = pd.read_excel(
                    file_path,
                    sheet_name=daily_sheet,
                    engine="openpyxl"
                )

                daily_df = daily_df.dropna(
                    how="all"
                )

                if not daily_df.empty:

                    self.dailyvariables_data.append(
                        daily_df
                    )

                    self.total_daily_rows += (
                        len(daily_df)
                    )

            else:

                self.log(
                    f"[WARNING] dailyvariables "
                    f"sheet missing in "
                    f"{file_path.name}"
                )

            self.processed_files += 1

            self.update_progress()

            self.update_statistics()

        except Exception as e:

            self.total_errors += 1

            self.log(
                f"[ERROR] Failed processing "
                f"{file_path.name}: {str(e)}"
            )

            self.update_statistics()

    # =====================================================
    # FORMAT EXCEL FILE
    # =====================================================

    def format_excel_file(self, file_path):

        try:

            wb = load_workbook(file_path)

            for ws in wb.worksheets:

                # HEADER STYLE
                for cell in ws[1]:

                    cell.font = Font(
                        bold=True,
                        color="FFFFFF"
                    )

                    cell.fill = PatternFill(
                        start_color="1F4E78",
                        end_color="1F4E78",
                        fill_type="solid"
                    )

                # AUTO WIDTH
                for column in ws.columns:

                    max_length = 0

                    column_letter = (
                        column[0].column_letter
                    )

                    for cell in column:

                        try:

                            if cell.value:

                                max_length = max(
                                    max_length,
                                    len(
                                        str(cell.value)
                                    )
                                )

                        except:
                            pass

                    adjusted_width = min(
                        max_length + 5,
                        60
                    )

                    ws.column_dimensions[
                        column_letter
                    ].width = adjusted_width

            wb.save(file_path)

        except Exception as e:

            self.log(
                f"[ERROR] Formatting failed: "
                f"{str(e)}"
            )

    # =====================================================
    # SAVE DAILYVARIABLES
    # =====================================================

    def save_dailyvariables(self):

        if not self.dailyvariables_data:

            self.log(
                "[WARNING] No dailyvariables data found"
            )

            return

        self.log(
            "[INFO] Saving dailyvariables..."
        )

        combined_df = pd.concat(
            self.dailyvariables_data,
            ignore_index=True
        )

        output_file = (
            self.output_dir
            / "all_dailyvariables.xlsx"
        )

        with pd.ExcelWriter(
            output_file,
            engine="openpyxl"
        ) as writer:

            combined_df.to_excel(
                writer,
                sheet_name="dailyvariables",
                index=False
            )

        self.format_excel_file(output_file)

        self.log(
            "[SUCCESS] Saved all_dailyvariables.xlsx"
        )

    # =====================================================
    # SAVE CHUNK FILES
    # =====================================================

    def save_chunk_files(self):

        self.log(
            "[INFO] Saving chunk files..."
        )

        equipment_df = pd.concat(
            self.equipment_data,
            ignore_index=True
        ) if self.equipment_data else pd.DataFrame()

        mpdm_df = pd.concat(
            self.mpdm_data,
            ignore_index=True
        ) if self.mpdm_data else pd.DataFrame()

        max_rows = max(
            len(equipment_df),
            len(mpdm_df)
        )

        if max_rows == 0:

            self.log(
                "[WARNING] No equipment/mpdm data found"
            )

            return

        total_chunks = ceil(
            max_rows / self.chunk_size
        )

        for i in range(total_chunks):

            start = i * self.chunk_size

            end = start + self.chunk_size

            eq_chunk = equipment_df.iloc[
                start:end
            ]

            mpdm_chunk = mpdm_df.iloc[
                start:end
            ]

            output_file = (
                self.output_dir
                / f"dataset_chunk_{i+1}.xlsx"
            )

            with pd.ExcelWriter(
                output_file,
                engine="openpyxl"
            ) as writer:

                eq_chunk.to_excel(
                    writer,
                    sheet_name="equipment",
                    index=False
                )

                mpdm_chunk.to_excel(
                    writer,
                    sheet_name="mpdm",
                    index=False
                )

            self.format_excel_file(output_file)

            self.log(
                f"[SUCCESS] Saved: "
                f"{output_file.name}"
            )

    # =====================================================
    # MAIN PROCESS
    # =====================================================

    def process(self):

        self.log(
            "[INFO] Starting processing..."
        )

        excel_files = list(
            self.input_dir.glob("*.xlsx")
        )

        self.total_files = len(excel_files)

        if self.total_files == 0:

            raise Exception(
                "No Excel files found."
            )

        self.log(
            f"[INFO] Found "
            f"{self.total_files} Excel files"
        )

        # MULTI-THREADING
        with ThreadPoolExecutor(
            max_workers=self.max_workers
        ) as executor:

            executor.map(
                self.process_single_file,
                excel_files
            )

        # SAVE OUTPUTS
        self.save_dailyvariables()

        self.save_chunk_files()

        self.log(
            "[SUCCESS] Processing completed."
        )

        self.update_progress()

        self.update_statistics()